# coding: utf-8
import pandas as pd
import numpy as np
import sys
import os
import logging
import cx_Oracle
from openpyxl import load_workbook
reload(sys)
sys.setdefaultencoding('utf-8')
from configparser import ConfigParser

class logger():
    def __init__(self, path):
        self.path = path
        self.logger = self.__pz(self.path)

    def __pz(self, path):
        path1 = path + u'\log.log'
        logging.basicConfig(level=logging.INFO,
                            format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                            filename='%s' % path1,
                            filemode='w+')
        # create logger
        logger_name = "hangBed_sevendaysNonessential"
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.INFO)

        # create file handler
        fh = logging.StreamHandler()
        fh.setLevel(logging.INFO)

        # create formatter
        fmt = "%(asctime)-15s %(levelname)s %(filename)s %(lineno)d  %(message)s"
        datefmt = "%a %d %b %Y %H:%M:%S"
        formatter = logging.Formatter(fmt, datefmt)
        # add handler and formatter to logger
        fh.setFormatter(formatter)
        logging.getLogger('').addHandler(fh)
        return logger



def readData(rawData):
    rawData1 = rawData.rename(columns={u'分类名称':'category',u'医院编码':'hosid',u'就诊登记号':'clinid',u'药品剂型':'drugForm',u'医院所在分中心':'hosCenter',u'医院名称':'hosname'})
    # 去掉时间错误的
    rawData1['year'] = rawData1[u'明细发生时间'].str.extract("(\d{4})").astype(int)
    rawData2 = rawData1[(rawData1['year'] < 2020) & (rawData1['year'] > 2000)]
    rawData2['feetime'] = rawData2[u'明细发生时间'].apply(lambda x: pd.to_datetime(x).date())
    rawData2['intime'] = rawData2[u'入院时间'].apply(lambda x: pd.to_datetime(x).date())
    data = rawData2[['clinid','hosid','intime','feetime','category','drugForm']]
    #根据医院、就诊id、入院时间排序
    data = data.sort_values(by=['hosid','clinid','intime'])
    return data


#无实质性治疗，只有床位、护理、其他诊疗
def flag(lists):
    if lists.isin([u'床位',u'护理',u'其它诊疗']).all():
        return 1
    else:
        return 0

def findMaxConsecutiveOnes(nums):
    cnt = 0
    ans = 0
    for i in nums:
        if i ==1:
            cnt = cnt +1
            ans = max(ans, cnt)
        else:
            cnt = 0
    return ans

def mkdir(path,log):
    # 引入模块
    isExists = os.path.exists(path)
    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        log.info('%s directory create success' % path)
        #print path + ' 创建成功'
        # 创建目录操作函数
        os.makedirs(path)
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        log.info('%s directory already exists' % path)
        #print path + ' 目录已存在'
        return False


if __name__=='__main__':
    # TODO 获取当前目录
    path = os.path.split(os.path.realpath(__file__))[0]
    path = path.decode('gbk')
    logger1 = logger(path)
    log = logger1.logger

    # TODO 开始运行
    log.info('run begin')

    # TODO 创建结果目录
    resultPath = path + u'\\result'
    mkdir(resultPath, log)

    # TODO 读取配置文件
    log.info('Read configuration file begin')
    cf = ConfigParser()
    Locpath = path + u'\\Config.ini'
    cf.read(Locpath)
    category = cf.get('parameter', 'category')
    days = cf.getint('parameter', 'days')
    log.info('Read configuration file end')

    # TODO 读取数据
    log.info('Read data begin')
    datapath = path + u'\\data'
    listDir = os.listdir(datapath)
    for i in range(len(listDir)):
        filePath = os.path.join(datapath, listDir[i])
        if os.path.isfile(filePath):
            if '明细' in listDir[i]:
                rawData = pd.read_csv(filePath, encoding='gb18030',
                                      dtype={u'药品剂型': np.str, u'分类名称': np.str, u'中心端名称': np.str, u'就诊登记号': np.str})
            else:
                settleData = pd.read_csv(filePath, encoding='gb18030', dtype={u'就诊登记号': np.str})
    log.info('Read rawData end')

    # TODO 开始分析
    data = readData(rawData)
    data = data.fillna('nan')
    log.info('Read Data end')

    #某天的明细是否只包括床位、护理、其他诊疗，是取1，非取0
    df = data['category'].groupby([data['clinid'],data['feetime']]).apply(flag).reset_index()
    df.rename(columns={'category':'flag'},inplace=True)

    # 查看连续7天以上没有发生实质性治疗的
    df_1 = df['flag'].groupby(df['clinid']).apply(findMaxConsecutiveOnes).reset_index()
    df_2 = df_1[df_1['flag'] > days]

    #输出结果
    idlist = []
    categoryPath = os.path.join(resultPath, category)
    for i in list(df_2['clinid']):
        output = rawData[rawData[u'就诊登记号'] == str(i)]
        outputdata = pd.merge(output, settleData,left_on=u'就诊登记号',right_on=u'就诊登记号',how='left' )
        outputdata[u'报销金额'] = outputdata[u'报销金额'].astype(float)
        outputdata['claimRatio'] = (outputdata[u'报销金额'] / outputdata[u'医疗总费用']).astype(float)

        if list(outputdata['claimRatio'])[1]>0.5:
            idlist.append(i)
            districts = list(outputdata[u'医院所在分中心'])[1]
            hos = list(outputdata[u'医院名称_x'])[1]
            districtsPath = os.path.join(categoryPath, districts)
            mkdir(districtsPath, log)
            filename = districtsPath + '/' + hos + '.xlsx'
            if not os.path.exists(filename):
                writer = pd.ExcelWriter(filename)
                outputdata.to_excel(writer, sheet_name=i, encoding='gb18030')
                writer.save()
            else:
                book = load_workbook(filename)
                writer = pd.ExcelWriter(filename, engine='openpyxl')
                writer.book = book
                outputdata.to_excel(writer, sheet_name=i, encoding='gb18030')
                writer.save()

    df_3 = df_2[df_2.clinid.isin(idlist)]
    flagname = resultPath + '/flag_'+category+'.xlsx'
    if not os.path.exists(flagname):
        df_3.to_excel(flagname, encoding='gb18030')
    else:
        book = load_workbook(flagname)
        writer = pd.ExcelWriter(flagname, engine='openpyxl')
        writer.book = book
        df_3.to_excel(writer, encoding='gb18030')
        writer.save()
    log.info('run end')






